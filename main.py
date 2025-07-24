import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
import time
from datetime import datetime

# === CONFIGURAÇÕES ===
CAMINHO_EXCEL = r'C:\Scripts_e_Automacoes\MeuDanfe_NFe\Chaves_de_Acesso.xlsx'
PASTA_DESTINO = r'C:\Scripts_e_Automacoes\MeuDanfe_NFe\XML'
PASTA_LOGS = r'C:\Scripts_e_Automacoes\MeuDanfe_NFe\Logs'
URL_SITE = 'https://meudanfe.com.br'
TEMPO_TIMEOUT = 120  # segundos

# === LOG ===
def criar_arquivo_log():
    if not os.path.exists(PASTA_LOGS):
        os.makedirs(PASTA_LOGS)
    nome_arquivo = datetime.now().strftime("log_%d-%m-%Y_%H-%M.txt")
    caminho_log = os.path.join(PASTA_LOGS, nome_arquivo)
    return open(caminho_log, "w", encoding="utf-8")

def escrever_log(log_file, mensagem):
    agora = datetime.now().strftime("%H:%M:%S")
    log_file.write(f"[{agora}] {mensagem}\n")
    print(f"[{agora}] {mensagem}")

# === BROWSER ===
def configurar_chromedriver():
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    prefs = {
        "safebrowsing.enabled": True,
        "download.default_directory": PASTA_DESTINO,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True
    }
    options.add_experimental_option("prefs", prefs)
    return uc.Chrome(options=options, use_subprocess=True)

# === PLANILHA ===
def carregar_planilha():
    wb = openpyxl.load_workbook(CAMINHO_EXCEL)
    sheet = wb.active
    return wb, sheet

# === DOWNLOAD XML ===
def baixar_xml(driver, wait, chave, row, log_file):
    try:
        escrever_log(log_file, f"🔎 Buscando chave: {chave}")

        if row[1].value and row[1].value.strip() != "":
            escrever_log(log_file, f"⏭️  Chave {chave} já possui status: {row[1].value}. Pulando.")
            return driver, wait

        nome_arquivo_esperado = os.path.join(PASTA_DESTINO, f"{chave}.xml")
        nome_com_prefixo = os.path.join(PASTA_DESTINO, f"NFE-{chave}.xml")
        if os.path.exists(nome_arquivo_esperado) or os.path.exists(nome_com_prefixo):
            escrever_log(log_file, f"⏩ XML já existe localmente para a chave {chave}.")
            row[1].value = "SUCESSO"
            return driver, wait

        arquivos_antes = set(os.listdir(PASTA_DESTINO))

        # Preenche e clica em BUSCAR
        search_box = wait.until(EC.element_to_be_clickable((By.ID, 'searchTxt')))
        search_box.clear()
        driver.execute_script(f"document.getElementById('searchTxt').value = '{chave}';")
        driver.find_element(By.ID, 'searchBtn').click()
        time.sleep(2)

        # Espera até 3 minutos
        timeout_inicio = time.time()
        while time.time() - timeout_inicio < 180:
            arquivos_depois = set(os.listdir(PASTA_DESTINO))
            novos = arquivos_depois - arquivos_antes
            if novos:
                novo_arquivo = novos.pop()
                caminho_origem = os.path.join(PASTA_DESTINO, novo_arquivo)
                caminho_destino = os.path.join(PASTA_DESTINO, f"{chave}.xml")

                while novo_arquivo.endswith(".crdownload") or not os.path.exists(caminho_origem):
                    time.sleep(1)
                    arquivos_depois = set(os.listdir(PASTA_DESTINO))
                    novos = arquivos_depois - arquivos_antes
                    if novos:
                        novo_arquivo = novos.pop()
                        caminho_origem = os.path.join(PASTA_DESTINO, novo_arquivo)

                os.rename(caminho_origem, caminho_destino)
                row[1].value = "SUCESSO"
                escrever_log(log_file, f"✅ XML baixado automaticamente: {chave}")
                return driver, wait

            # Tenta clicar no botão "Baixar XML"
            try:
                btn_baixar = driver.find_element(By.ID, 'downloadXmlBtn')
                if btn_baixar.is_displayed() and btn_baixar.is_enabled():
                    btn_baixar.click()
                    escrever_log(log_file, f"⬇️ Clicou no botão 'Baixar XML'")
                    time.sleep(2)
            except:
                pass

            time.sleep(2)

        # Timeout de 3 minutos
        row[1].value = "FALHA"
        escrever_log(log_file, f"❌ Timeout aguardando download da chave: {chave}. Reiniciando navegador.")

        driver.quit()
        driver = configurar_chromedriver()
        driver.get(URL_SITE)
        wait = WebDriverWait(driver, TEMPO_TIMEOUT)
        time.sleep(2)

    except Exception as e:
        escrever_log(log_file, f"❌ Erro ao processar chave {chave}: {e}")
        row[1].value = "ERRO"

    return driver, wait

# === MAIN ===
def main():
    if not os.path.exists(PASTA_DESTINO):
        os.makedirs(PASTA_DESTINO)
    log_file = criar_arquivo_log()

    driver = configurar_chromedriver()
    driver.get(URL_SITE)
    wait = WebDriverWait(driver, TEMPO_TIMEOUT)
    time.sleep(3)

    wb, sheet = carregar_planilha()

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        chave = row[0].value
        if chave:
            driver, wait = baixar_xml(driver, wait, str(chave).strip(), row, log_file)

    try:
        wb.save(CAMINHO_EXCEL)
        escrever_log(log_file, "📄 Planilha atualizada com sucesso.")
    except Exception as e:
        escrever_log(log_file, f"⚠️ Erro ao salvar planilha: {e}")
    finally:
        wb.close()
        try:
            driver.quit()
        except Exception as e:
            escrever_log(log_file, f"⚠️ Erro ao fechar navegador: {e}")
        log_file.close()

if __name__ == "__main__":
    main()
