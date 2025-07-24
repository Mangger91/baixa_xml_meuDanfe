import os
import openpyxl

caminho_planilha = r'C:\Scripts_e_Automacoes\MeuDanfe_NFe\Chaves_de_Acesso.xlsx'
pasta_xml = r'C:\Scripts_e_Automacoes\MeuDanfe_NFe\XML'

def validar_downloads():
    wb = openpyxl.load_workbook(caminho_planilha)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):
        chave_cell = row[0]
        status_cell = row[1] if len(row) > 1 else None
        
        chave = str(chave_cell.value).strip()
        if not chave:
            continue
        
        if status_cell and status_cell.value:
            continue
        
        caminho_xml = os.path.join(pasta_xml, f"NFE-{chave}.xml")
        
        status = "SUCESSO" if os.path.exists(caminho_xml) else "FALHA"
        
        sheet.cell(row=chave_cell.row, column=2, value=status)
        print(f"{chave}: {status}")
               
    wb.save(caminho_planilha)
    wb.close()
    print("Validação concluída. Status atualizado na planilha.")
            
if __name__ == "__main__":
    validar_downloads()